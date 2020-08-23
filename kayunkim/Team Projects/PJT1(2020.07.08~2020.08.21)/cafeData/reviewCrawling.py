import os
from time import sleep

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup

from openpyxl import load_workbook

##############################################################  ############
##################### variable related selenium ##########################
##########################################################################
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('lang=ko_KR')
chromedriver_path = "chromedriver"
driver = webdriver.Chrome(os.path.join(os.getcwd(), chromedriver_path), options=options)  # chromedriver 열기


##########################################################################
###################### variable related excel ############################
##########################################################################
load_file_name = "cafeInfoLists.xlsx"
load_wb = load_workbook(os.path.join(os.getcwd(), load_file_name))
load_ws = load_wb['Sheet']
get_cells = load_ws['A6807':'C10000']

review_file_name = "reviewLists.xlsx"
review_wb = load_workbook(os.path.join(os.getcwd(), review_file_name))
review_ws = review_wb['Sheet1']

row_num = 4165 # reviewLists.xslx에 쓸 행 번호

def main():
    global driver, load_wb

    driver.implicitly_wait(4)  # 렌더링 될때까지 기다린다 4초
    driver.get('https://map.kakao.com/')  # 주소 가져오기

    # 카페 정보 읽어오기
    place_infos = []
    for rows in get_cells:
        place = []
        for cell in rows:
            place.append(cell.value)
        place_infos.append(place)
    print('get places')

    for i, place in enumerate(place_infos):
        if i % 4 == 0 and i != 0:
            sleep(5)
            review_wb.save(os.path.join(os.getcwd(), review_file_name))
        print("#####", i)
        search(place)

    driver.quit()
    print("finish")

def search(place_info):
    global driver
    cafe_id = place_info[0]
    place = place_info[1]
    address = place_info[2]

    search_area = driver.find_element_by_xpath('//*[@id="search.keyword.query"]') # 검색 창
    search_area.send_keys(place)  # 검색어 입력
    driver.find_element_by_xpath('//*[@id="search.keyword.submit"]').send_keys(Keys.ENTER)  # Enter로 검색
    sleep(1)

    # 검색된 정보가 있는 경우에만 탐색
    # 1번 페이지 cafe list 읽기
    html = driver.page_source

    soup = BeautifulSoup(html, 'html.parser')
    cafe_lists = soup.select('.placelist > .PlaceItem')

    if crawling(place, address, cafe_id, cafe_lists):
        search_area.clear()
    else:
        # 페이지 내에 일치하는게 없을 경우 다음 페이지도 찾아보기
        # 우선 더보기 클릭해서 2페이지
        try:
            driver.find_element_by_xpath('//*[@id="info.search.place.more"]').send_keys(Keys.ENTER)
            sleep(1)

            # 2~ 5페이지 읽기
            # 여기도 없으면 없는걸로 간주?
            for i in range(2, 6):
                # 페이지 넘기기
                xPath = '//*[@id="info.search.page.no' + str(i) + '"]'
                driver.find_element_by_xpath(xPath).send_keys(Keys.ENTER)
                sleep(1)

                html = driver.page_source
                soup = BeautifulSoup(html, 'html.parser')
                cafe_lists = soup.select('.placelist > .PlaceItem')

                # 일치하는거 찾았으면 종료 아니면 다음 페이지 찾기기
                if crawling(place, address, cafe_id, cafe_lists):
                    return
        except ElementNotInteractableException:
            print('not found')
        finally:
            search_area.clear()

def crawling(place, address, cafe_id, cafe_lists):
    global row_num, review_wb, review_ws
    """
    페이지 목록을 받아서 크롤링 하는 함수
    :param place: 리뷰 정보 찾을 카페이름
    :param address: 리뷰 정보 찾을 카페 주소
    :param cafe_lists: 검색했을때 나타난 검색된 목록
    :return: 원하는 카페를 찾았는지 여부
    """

    flag = False
    for i, cafe in enumerate(cafe_lists):
        cafe_name = cafe.select('.head_item > .tit_name > .link_name')[0].text  # cafeName
        cafe_address = cafe.select('.info_item > .addr > p')[0].text  # cafe address

        # 카페 정보가 일치하는 경우에만 상세정보 페이지에서 리뷰 읽기
        if cafe_name == place and cafe_address == address:
            detail_page_xpath = '//*[@id="info.search.place.list"]/li[' + str(i + 1) + ']/div[5]/div[4]/a[1]'
            try:
                driver.find_element_by_xpath(detail_page_xpath).send_keys(Keys.ENTER)
                driver.switch_to.window(driver.window_handles[-1])  # 상세정보 탭으로 변환
                sleep(1)

                html = driver.page_source
                soup = BeautifulSoup(html, 'html.parser')

                # 리뷰 목록 찾기
                review_lists = soup.select('.list_evaluation > li')

                if len(review_lists) != 0:
                    review_ws.cell(row_num, 1, cafe_id)    # 카페 아이디
                    review_ws.cell(row_num, 2, cafe_name)  # 카페 이름

                    for i, review in enumerate(review_lists):
                        comment = review.select('.txt_comment > span')

                        if len(comment) != 0:
                            # 저장 (cell, row, data)
                            review_ws.cell(row_num, i+3, comment[0].text)
                    row_num += 1
                else:
                    print('no review')
                driver.close()
                driver.switch_to.window(driver.window_handles[0])  # 검색 탭으로 전환
                return True
            except NoSuchElementException:
                print("not cafe info list")


    return flag
if __name__ == "__main__":
    main()
