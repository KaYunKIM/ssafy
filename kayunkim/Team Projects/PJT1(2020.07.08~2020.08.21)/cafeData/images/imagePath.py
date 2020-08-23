import os
import random
from openpyxl import load_workbook

saveCafeInfoFileName = "outputcafelist.xlsx"
write_wb = load_workbook(os.path.join(os.getcwd(), saveCafeInfoFileName))
write_ws = write_wb['outputcafelist']

savePathFileName = 'pathlist.xlsx'
save_wb = load_workbook(os.path.join(os.getcwd(), savePathFileName))
save_ws = save_wb['Sheet1']
def main():
    global write_ws

    get_cells = write_ws['F2':'F27053']
    default_path = '/home/data/thumb_img/default'
    for j, rows in enumerate(get_cells):
        for cell in rows:
            path = cell.value
            if path is None:
                x = random.randint(1,10)
                save_path = default_path + str(x) + '.jpg'
                print(save_path)
                save_ws.cell(j+2, 1, save_path)

    save_wb.save(os.path.join(os.getcwd(), savePathFileName))

if __name__ == "__main__":
    main()